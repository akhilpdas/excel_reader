from django.shortcuts import render
from django.views.generic.base import View
from .forms import *
from .models import *
from django.views.generic import ListView, DetailView
import openpyxl


# Create your views here.


class UploadView(View):
    """Uploading excel file."""

    template_name = 'excel_info/upload.html'
    form_class = UploadFileForm

    def get(self, request):
        """Get."""

        form = self.form_class
        context = {
            'form': form,
        }
        return render(request, self.template_name, context)
    
    def post(self, request):
        """Post."""

        form = self.form_class(request.POST, request.FILES)
        print("form", form)
        
        if form.is_valid():
            """Process the excel file."""
            file = request.FILES['file']
            extention = str(file).split(".")[-1]
            print("extention", extention)
            if extention not in ['xlsx', 'xls', 'xlsm']:
                context = {
                    'form': form,
                    'status': '1',
                }
                return render(request, self.template_name, context)

            try:
                """Checking the file already uploaded."""
                info = ExcelInfo.objects.get(file_name=file)
                if bool(info):
                    print("already existing")
                    context = {
                        'form': form,
                        'status': '2',
                    }
                    return render(request, self.template_name, context)

            except:
                pass
            print("file", file)
            wb = openpyxl.load_workbook(file) 
            sheet = wb.worksheets[0]

            max_rows = sheet.max_row
            max_columns =  sheet.max_column

            data = {}
            print("max", max_rows, max_columns)

            """Reading data from excel sheet."""
            for col_no in range(1, max_columns+1):
                print("cols")
                heading = True
                item_list = []
                for row_no in range(1, max_rows+1):
                    print("row")
                    if heading:
                        key = str(sheet.cell(row=row_no, column=col_no).value)
                        data[key] = []
                        heading = False
                    else:
                        item_list.append(str(sheet.cell(row=row_no, column=col_no).value))
                data[key] = item_list
            print("data", data)

            ExcelInfo.objects.create(file_name=file,
                                     data=data)
            return http.HttpResponseRedirect(reverse('FileListView'))
        else:

            context = {
                'form': form,
            }
            return render(request, self.template_name, context)


class FileListView(ListView):
    """Showing files."""

    template_name = 'excel_info/file_list.html'
    context_object_name = 'file_list'
    model = ExcelInfo
    


class FileDetailsView(DetailView):
    """Showing details of a file."""
     
    template_name = 'excel_info/file_details.html'
    model = ExcelInfo
    pk_url_kwarg = 'pk'
    context_object_name = 'file_detail'
